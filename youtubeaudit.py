"""
YouTube Channel Audit Tool
──────────────────────────
Run this script on your local machine.

Requirements:
    pip install requests openpyxl

Usage:
    python youtube_audit.py

Output:
    youtube_channel_audit.xlsx  (in the same folder)
"""

import os
import requests
from dotenv import load_dotenv
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

load_dotenv()

# ── CONFIG ──────────────────────────────────────────────────────────
API_KEY = os.getenv("YOUTUBE_API_KEY")
if not API_KEY:
    raise RuntimeError("YOUTUBE_API_KEY is not set. Add it to .env or your environment variables.")
OUTPUT_FILE = "youtube_channel_audit.xlsx"

# ── CHANNELS ─────────────────────────────────────────────────────────
# Format: (display_name, id_type, identifier)
# id_type = "handle" for @handles, "id" for channel IDs
CHANNELS = [
    ("Admire1975",               "handle", "Admire1975"),
    ("Prerna Sharma",            "handle", "prernasharma1515"),
    ("Bhavya Collections",       "handle", "bhavyacollections"),
    ("Eeja Apparels",            "handle", "EejaApparels"),
    ("Nav Online Creation",      "handle", "nav__online__creation"),
    ("Lord Shiv Collection",     "handle", "lordshivcollection515"),
    ("The Unique Creations BLR", "handle", "the_unique_creations_blr"),
    ("PS Fashion Hub",           "handle", "Psfa.shionhub13"),
    ("Channel UCPqEe",           "id",     "UCPqEe-BSedYN8XUqVWeSizg"),
    ("Channel UCacQO",           "id",     "UCacQO0wZGlJvJ2oAauCv6DQ"),
    ("Channel UC7TED",           "id",     "UC7TEDmxM33CdMFe6fKpNfMA"),
    ("Channel UCS0qX",           "id",     "UCS0qX6OL_xXoezKgtY542zQ"),
    ("Channel UC5N1D",           "id",     "UC5N1DBOilQzS02YCctBlIXw"),
    ("Channel UCBPGZ",           "id",     "UCBPGzCGP-PL0-uGGoLMaHwQ"),
    ("Channel UCyWuy",           "id",     "UCyWuybNVL1Yu-yIMUxrAWWg"),
    ("Channel UCEjv0",           "id",     "UCEjv0SIE0PY34vtSKImYF5Q"),
    ("Channel UCQ_Jh",           "id",     "UCQ_JhurmM1NmrLBuHljqplw"),
    ("Channel UCaQRL",           "id",     "UCaQRLu8KM3ednHvJQdOtLyg"),
    ("Kavi Collection",          "id",     "UCl8sfla2MB09tgyTRo6RS9A"),
    ("New Channel",              "id",     "UCIGo5g2SqBaysj0k4GD92mA"),
]

# ── SEO KEYWORDS ──────────────────────────────────────────────────────
SEO_KEYWORDS = [
    "saree", "kurti", "collection", "shop", "buy", "wholesale", "retail",
    "fashion", "apparel", "clothing", "ethnic", "online", "india", "dress",
    "silk", "cotton", "lehenga", "dupatta", "suit", "blouse", "wear",
    "kurta", "salwar", "banarasi", "chiffon", "georgette", "handloom"
]

# ── PROFILE PIC CHECK CRITERIA ───────────────────────────────────────
PROFILE_PIC_CRITERIA = {
    "face":    "Manual check needed",   # Cannot be determined via API
    "yellow":  "Manual check needed",   # Cannot be determined via API
    "outline": "Manual check needed",   # Cannot be determined via API
}

# ─────────────────────────────────────────────────────────────────────
# API HELPERS
# ─────────────────────────────────────────────────────────────────────

BASE = "https://www.googleapis.com/youtube/v3"

def api_get(endpoint, params):
    params["key"] = API_KEY
    r = requests.get(f"{BASE}/{endpoint}", params=params, timeout=10)
    r.raise_for_status()
    return r.json()

def fetch_channel(id_type, identifier):
    if id_type == "handle":
        return api_get("channels", {
            "part": "snippet,brandingSettings,contentDetails",
            "forHandle": identifier
        })
    else:
        return api_get("channels", {
            "part": "snippet,brandingSettings,contentDetails",
            "id": identifier
        })

def fetch_latest_video(uploads_playlist_id):
    data = api_get("playlistItems", {
        "part": "snippet",
        "playlistId": uploads_playlist_id,
        "maxResults": 1
    })
    items = data.get("items", [])
    if not items:
        return None, None
    video_id = items[0]["snippet"]["resourceId"]["videoId"]
    vdata = api_get("videos", {"part": "snippet", "id": video_id})
    vitems = vdata.get("items", [])
    if not vitems:
        return video_id, None
    return video_id, vitems[0]["snippet"]

# ─────────────────────────────────────────────────────────────────────
# ANALYSIS HELPERS
# ─────────────────────────────────────────────────────────────────────

def analyze_seo(text):
    """Returns (Yes/No/Partial, notes_string)"""
    if not text or not text.strip():
        return "No", "No text provided"
    has_hashtags = "#" in text
    found_kw = [k for k in SEO_KEYWORDS if k.lower() in text.lower()]
    if has_hashtags and len(found_kw) >= 2:
        return "Yes", f"Hashtags ✓ | Keywords: {', '.join(found_kw[:5])}"
    elif has_hashtags or len(found_kw) >= 2:
        tag_note = "Hashtags ✓" if has_hashtags else "No hashtags"
        kw_note = f"Keywords: {', '.join(found_kw[:5])}" if found_kw else "No relevant keywords"
        return "Partial", f"{tag_note} | {kw_note}"
    else:
        return "No", "No hashtags or relevant fashion keywords found"

def banner_status(branding):
    url = branding.get("image", {}).get("bannerExternalUrl", "")
    if url:
        return "Yes", url
    return "No", "No banner image set"

# ─────────────────────────────────────────────────────────────────────
# MAIN FETCH LOOP
# ─────────────────────────────────────────────────────────────────────

def fetch_all():
    results = []
    for display_name, id_type, identifier in CHANNELS:
        print(f"  Fetching: {display_name}...", end=" ", flush=True)
        try:
            data = fetch_channel(id_type, identifier)
            items = data.get("items", [])
            if not items:
                raise ValueError("Channel not found via API")

            item      = items[0]
            channel_id = item["id"]
            snippet   = item.get("snippet", {})
            branding  = item.get("brandingSettings", {})
            cd        = item.get("contentDetails", {})

            title       = snippet.get("title", display_name)
            description = snippet.get("description", "")
            pic_url     = snippet.get("thumbnails", {}).get("high", {}).get("url", "")
            uploads_id  = cd.get("relatedPlaylists", {}).get("uploads", "")

            # Description
            desc_exists = "Yes" if description.strip() else "No"
            desc_seo, desc_notes = analyze_seo(description)
            desc_preview = (description[:150] + "...") if len(description) > 150 else description

            # Banner
            ban_exists, ban_notes = banner_status(branding)
            ban_good = "Manual check needed" if ban_exists == "Yes" else "N/A"

            # Home / latest video
            hv_exists = hv_seo = hv_notes = "N/A"
            if uploads_id:
                vid_id, vsnippet = fetch_latest_video(uploads_id)
                if vsnippet:
                    hv_exists = "Yes"
                    vtitle    = vsnippet.get("title", "")
                    vdesc     = vsnippet.get("description", "")
                    hv_seo, hv_seo_notes = analyze_seo(vdesc)
                    hv_notes  = f'Latest: "{vtitle[:60]}" | {hv_seo_notes}'
                elif vid_id:
                    hv_exists = "Yes (no desc fetched)"
                    hv_seo    = "Unknown"
                    hv_notes  = ""
                else:
                    hv_exists = "No"
                    hv_seo    = "N/A"

            results.append({
                "name":        title,
                "url":         f"https://www.youtube.com/channel/{channel_id}",
                "pic_url":     pic_url,
                "desc_exists": desc_exists,
                "desc_seo":    desc_seo,
                "desc_notes":  desc_notes,
                "desc_preview":desc_preview,
                "hv_exists":   hv_exists,
                "hv_seo":      hv_seo,
                "hv_notes":    hv_notes,
                "ban_exists":  ban_exists,
                "ban_good":    ban_good,
                "ban_notes":   ban_notes,
                "error":       None
            })
            print("✓")

        except Exception as e:
            ch_url = (f"https://www.youtube.com/@{identifier}"
                      if id_type == "handle"
                      else f"https://www.youtube.com/channel/{identifier}")
            results.append({
                "name": display_name, "url": ch_url,
                "pic_url": "", "desc_exists": "Error", "desc_seo": "Error",
                "desc_notes": str(e)[:100], "desc_preview": "",
                "hv_exists": "Error", "hv_seo": "Error", "hv_notes": "",
                "ban_exists": "Error", "ban_good": "Error", "ban_notes": "",
                "error": str(e)
            })
            print(f"✗  {e}")

    return results

# ─────────────────────────────────────────────────────────────────────
# EXCEL BUILDER
# ─────────────────────────────────────────────────────────────────────

def s(c): return Side(style="thin", color=c)
def bdr(c="C8D0DC"): return Border(left=s(c), right=s(c), top=s(c), bottom=s(c))
def bg(hex): return PatternFill("solid", fgColor=hex)

# Value → cell colour
STATUS_COLORS = {
    "Yes":      "D6F4D0",   # green
    "No":       "FDDEDE",   # red
    "Partial":  "FFF3CD",   # amber
    "N/A":      "F0F0F0",   # grey
    "Error":    "FFD0D0",   # pink
    "Manual check needed": "E8F4FD",  # light blue
    "Unknown":  "F0F0F0",
}

def status_fill(val):
    for k, color in STATUS_COLORS.items():
        if val and val.startswith(k):
            return bg(color)
    return bg("FFFFFF")

def build_excel(results):
    wb = Workbook()
    ws = wb.active
    ws.title = "YouTube Channel Audit"
    ws.sheet_properties.tabColor = "1565C0"

    # ── Section header (row 1) ─────────────────────────────────────
    sections = [
        (1, 3,  "CHANNEL INFO",            "2C3E50"),
        (4, 7,  "🖼  PROFILE PICTURE",      "1565C0"),
        (8, 10, "📝  CHANNEL DESCRIPTION",  "2E7D32"),
        (11,13, "🎬  LATEST / HOME VIDEO",  "6A1B9A"),
        (14,16, "🖼  BANNER",               "AD1457"),
        (17,17, "📋  OVERALL NOTES",        "37474F"),
    ]
    ws.row_dimensions[1].height = 28
    for (sc, ec, label, color) in sections:
        cell = ws.cell(row=1, column=sc, value=label)
        cell.font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
        cell.fill = bg(color)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = bdr()
        if ec > sc:
            ws.merge_cells(start_row=1, start_column=sc, end_row=1, end_column=ec)

    # ── Column headers (row 2) ─────────────────────────────────────
    headers = [
        ("#",                           5),
        ("Channel Name",               22),
        ("Channel URL",                38),
        # Profile pic
        ("Face Shown?\n(Manual)",       16),
        ("Yellow BG?\n(Manual)",        16),
        ("Blue Outline?\n(Manual)",     17),
        ("Profile Pic URL",             36),
        # Description
        ("Description\nExists?",        14),
        ("Description\nSEO Optimised?", 18),
        ("Description Notes",           36),
        # Home video
        ("Home Video\nExists?",         14),
        ("Home Video\nSEO Optimised?",  18),
        ("Home Video Notes",            42),
        # Banner
        ("Banner\nExists?",             13),
        ("Banner\nQuality?",            18),
        ("Banner URL / Notes",          42),
        # Overall
        ("Overall Notes",               36),
    ]
    ws.row_dimensions[2].height = 40
    for col_idx, (label, width) in enumerate(headers, 1):
        c = ws.cell(row=2, column=col_idx, value=label)
        c.font = Font(name="Arial", size=9, bold=True, color="FFFFFF")
        c.fill = bg("1A1A2E")
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = bdr()
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # ── Data rows ─────────────────────────────────────────────────
    for i, r in enumerate(results):
        row = i + 3
        ws.row_dimensions[row].height = 28
        row_bg = "F7F9FC" if i % 2 == 0 else "FFFFFF"

        def wc(col, value, status=False, wrap=False, url=None):
            cell = ws.cell(row=row, column=col, value=value)
            cell.fill = status_fill(str(value)) if status else bg(row_bg)
            cell.border = bdr()
            cell.alignment = Alignment(horizontal="center" if status else "left",
                                       vertical="center", wrap_text=wrap)
            cell.font = Font(name="Arial", size=9,
                             color="1565C0" if url else "1A1A2E",
                             underline="single" if url else None)
            if url:
                cell.hyperlink = url
            return cell

        wc(1,  i + 1)
        wc(2,  r["name"])
        wc(3,  r["url"], url=r["url"])

        # Profile pic - manual cols
        wc(4,  "Manual check needed", status=True)
        wc(5,  "Manual check needed", status=True)
        wc(6,  "Manual check needed", status=True)
        if r["pic_url"]:
            pic_cell = wc(7, r["pic_url"], url=r["pic_url"])
        else:
            wc(7, "No profile picture URL found")

        # Description
        wc(8,  r["desc_exists"], status=True)
        wc(9,  r["desc_seo"],    status=True)
        wc(10, r["desc_notes"],  wrap=True)

        # Home video
        wc(11, r["hv_exists"], status=True)
        wc(12, r["hv_seo"],    status=True)
        wc(13, r["hv_notes"],  wrap=True)

        # Banner
        wc(14, r["ban_exists"], status=True)
        wc(15, r["ban_good"],   status=True)
        wc(16, r["ban_notes"],  wrap=True)

        # Overall notes (blank – user fills)
        wc(17, "", wrap=True)

    # ── Legend ────────────────────────────────────────────────────
    legend_row = len(results) + 3
    ws.row_dimensions[legend_row].height = 80
    legend = ws.cell(row=legend_row, column=1, value=(
        "📌  HOW TO USE THIS SHEET\n"
        "• Colour coding:  🟢 Green = Yes   🔴 Red = No   🟡 Amber = Partial   🔵 Blue = Manual check needed\n"
        "• Profile Pic columns (4–6): Visit each channel URL and manually check the profile picture.\n"
        "  → Face shown? Does the picture clearly show the person's face?\n"
        "  → Yellow BG? Is the background yellow?\n"
        "  → Blue Outline? Is there a blue border/outline around the image?\n"
        "• Description SEO: Automatically checked for keywords + hashtags via YouTube API.\n"
        "• Home Video SEO: Checked against the most recently uploaded video's description.\n"
        "• Banner Quality: URL is fetched automatically. Open the URL to visually evaluate quality.\n"
        "• Overall Notes: Use this column to add your own action items or priority flags per channel."
    ))
    legend.font = Font(name="Arial", size=9, italic=True, color="37474F")
    legend.fill = bg("EEF6FB")
    legend.alignment = Alignment(wrap_text=True, vertical="top")
    legend.border = bdr()
    ws.merge_cells(start_row=legend_row, start_column=1,
                   end_row=legend_row, end_column=17)

    # ── Freeze top 2 rows ─────────────────────────────────────────
    ws.freeze_panes = "A3"

    wb.save(OUTPUT_FILE)
    print(f"\n✅  Saved: {OUTPUT_FILE}")

# ─────────────────────────────────────────────────────────────────────
if __name__ == "__main__":
    print("🔍 Fetching YouTube channel data...\n")
    data = fetch_all()
    print(f"\n📊 Building Excel file...")
    build_excel(data)
    print("Done! Open youtube_channel_audit.xlsx to view results.")